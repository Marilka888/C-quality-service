"""
Produce a presentation-grade Excel report from the coverage pipeline.

Run against one or more parsed packages (data/packages/<id>.json format)
and emit a workbook with:

  - One sheet per package: requirements listed, colour-coded by status,
    with top-3 supporting PMI units inline.
  - A "Summary" sheet: per-package counts + package-level coverage score.
  - A "Findings" sheet: interesting gaps (MISSING requirements + low-
    confidence COVERED calls) hand-pickable as demo talking points.

The report uses the same "D" configuration we benchmarked last round:
classifier for requirement extraction, BoW retrieval, BGE cross-encoder
judge. On CPU ~3 min per package.

Usage:
    python scripts/generate_demo_report.py \
        --packages pkg_0002 pkg_0005 pkg_0008 pkg_0010 \
        --out ./demo_report.xlsx
"""
from __future__ import annotations

import argparse
import io
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, List

_THIS_DIR = Path(__file__).resolve().parent
_PROJECT_ROOT = _THIS_DIR.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.application.use_cases.build_coverage_units import CoverageUnitBuilder
from app.application.use_cases.build_requirements import RequirementBuilder
from app.application.use_cases.run_coverage_analysis import CoverageAnalysisPipeline
from app.core.config import CoverageConfig


# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------

# Colour palette: slightly desaturated so the report is easy on the eye
# during a demo projection.
_STATUS_FILL = {
    "COVERED":  PatternFill("solid", fgColor="C6E0B4"),   # green
    "PARTIAL":  PatternFill("solid", fgColor="FFE699"),   # yellow
    "MISSING":  PatternFill("solid", fgColor="F8CBAD"),   # orange
    "CONFLICT": PatternFill("solid", fgColor="F4B084"),   # red-orange
}
_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_WRAP = Alignment(vertical="top", wrap_text=True)


# ---------------------------------------------------------------------------
# Package → request conversion (same as other scripts)
# ---------------------------------------------------------------------------


def _chunks_to_prepared(pkg: dict) -> dict:
    documents = []
    for doc in pkg.get("documents", []):
        doc_type = (doc.get("doc_type") or "").upper()
        role = {"TZ": "tz", "PMI": "pmi", "PZ": "pz"}.get(doc_type)
        if not role:
            continue
        fragments = [
            {
                "fragment_id": c.get("chunk_id") or f"frag-{i}",
                "text": c.get("text") or "",
                "kind": "paragraph",
                "section_id": None,
            }
            for i, c in enumerate(doc.get("chunks", []))
            if (c.get("text") or "").strip()
        ]
        artifact = {
            "document_id": doc.get("doc_id") or doc_type,
            "package_id": pkg.get("package_id"),
            "doc_role": role,
            "sections": [],
            "fragments": fragments,
        }
        documents.append({
            "document_id": artifact["document_id"],
            "doc_role": role,
            "prepared_artifact": artifact,
        })
    return {
        "package_id": pkg.get("package_id"),
        "source_doc_role": "tz",
        "target_doc_roles": ["pmi", "pz"],
        "documents": documents,
        "options": {},
    }


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


_STATUS_ORDER = ["CONFLICT", "MISSING", "PARTIAL", "COVERED"]
_STATUS_SORT_KEY = {s: i for i, s in enumerate(_STATUS_ORDER)}


def _short_status(status) -> str:
    return str(status).rsplit(".", 1)[-1].rstrip("'>")


def _truncate(text: str, n: int = 300) -> str:
    text = (text or "").strip()
    if len(text) <= n:
        return text
    return text[: n - 1] + "…"


def _write_package_sheet(
    ws,
    pkg_name: str,
    requirements,
    results,
    pair_judgments,
) -> None:
    reqs_by_id = {r.req_id: r for r in requirements}

    headers = [
        "#",
        "Status",
        "Requirement text",
        "Section",
        "Classifier p",
        "Top evidence 1 (PMI unit)",
        "Evidence 1 status / conf",
        "Top evidence 2",
        "Evidence 2 status / conf",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
    ws.freeze_panes = "A2"

    # Sort: most alarming first (CONFLICT → MISSING → PARTIAL → COVERED)
    rows = []
    for r in results:
        d = r.model_dump() if hasattr(r, "model_dump") else r
        short = _short_status(d.get("status"))
        rows.append((_STATUS_SORT_KEY.get(short, 99), d))
    rows.sort(key=lambda x: x[0])

    for i, (_, result) in enumerate(rows, start=2):
        req_id = result["req_id"]
        req = reqs_by_id.get(req_id)
        req_text = req.text if req else "<unknown>"
        section = (req.source_section_id or "") if req else ""
        classifier_p = (req.metadata.get("classifier_score") if req else None)
        status = _short_status(result.get("status"))

        # Top 2 evidence
        evidence = sorted(
            result.get("evidence") or [],
            key=lambda e: (e.get("judgment") or {}).get("llm_confidence") or 0,
            reverse=True,
        )[:2]
        ev_texts = [""] * 2
        ev_summaries = [""] * 2
        for j, e in enumerate(evidence):
            ev_texts[j] = _truncate(e.get("text") or "", 220)
            jud = e.get("judgment") or {}
            label = _short_status(jud.get("rule_adjusted_label") or jud.get("llm_label") or "?")
            conf = jud.get("llm_confidence") or 0
            retr = e.get("retrieval_score") or 0
            ev_summaries[j] = f"{label} · conf={conf:.2f} · retr={retr:.2f}"

        row = [
            i - 1,
            status,
            _truncate(req_text, 400),
            section,
            f"{classifier_p:.3f}" if classifier_p is not None else "",
            ev_texts[0], ev_summaries[0],
            ev_texts[1], ev_summaries[1],
        ]
        ws.append(row)
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=col)
            cell.alignment = _WRAP
        # Tint the status cell
        fill = _STATUS_FILL.get(status)
        if fill:
            ws.cell(row=i, column=2).fill = fill

    # Column widths (empirical)
    widths = [5, 11, 52, 14, 12, 46, 22, 46, 22]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def _write_summary_sheet(ws, per_package: Dict[str, Dict]) -> None:
    headers = [
        "Package",
        "Requirements",
        "COVERED",
        "PARTIAL",
        "MISSING",
        "CONFLICT",
        "Coverage score",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN

    totals = Counter()
    for pkg, stats in per_package.items():
        counts = stats["counts"]
        n = sum(counts.values())
        # Weighted coverage: COVERED=1, PARTIAL=0.5, MISSING/CONFLICT=0
        score = (counts.get("COVERED", 0) + 0.5 * counts.get("PARTIAL", 0)) / max(1, n)
        row = [
            pkg,
            n,
            counts.get("COVERED", 0),
            counts.get("PARTIAL", 0),
            counts.get("MISSING", 0),
            counts.get("CONFLICT", 0),
            f"{score:.1%}",
        ]
        ws.append(row)
        for status in ("COVERED", "PARTIAL", "MISSING", "CONFLICT"):
            totals[status] += counts.get(status, 0)

    # Totals row
    totals_n = sum(totals.values())
    overall_score = (totals["COVERED"] + 0.5 * totals["PARTIAL"]) / max(1, totals_n)
    ws.append([
        "TOTAL",
        totals_n,
        totals["COVERED"], totals["PARTIAL"],
        totals["MISSING"], totals["CONFLICT"],
        f"{overall_score:.1%}",
    ])
    for col in range(1, len(headers) + 1):
        ws.cell(row=ws.max_row, column=col).font = Font(bold=True)

    widths = [14, 14, 12, 12, 12, 12, 16]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def _write_findings_sheet(ws, per_package: Dict[str, Dict]) -> None:
    headers = ["Package", "Status", "Requirement", "Why it's a finding", "Top evidence"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN

    row_idx = 2
    for pkg, stats in per_package.items():
        for finding in stats["findings"][:10]:  # cap per package
            row = [pkg, finding["status"], finding["req"], finding["reason"], finding["evidence"]]
            ws.append(row)
            for col, _ in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col).alignment = _WRAP
            fill = _STATUS_FILL.get(finding["status"])
            if fill:
                ws.cell(row=row_idx, column=2).fill = fill
            row_idx += 1

    widths = [14, 11, 50, 44, 46]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def _collect_findings(results, requirements) -> List[dict]:
    """Return interesting rows for the 'Findings' tab.

    Rules:
      - CONFLICT → always include (with evidence snippet)
      - MISSING with a classifier_p ≥ 0.9 → real gaps worth showing
      - COVERED with evidence confidence < 0.5 → "weak matches"
    """
    reqs_by_id = {r.req_id: r for r in requirements}
    out = []
    for r in results:
        d = r.model_dump() if hasattr(r, "model_dump") else r
        status = _short_status(d.get("status"))
        req = reqs_by_id.get(d["req_id"])
        req_text = _truncate(req.text, 200) if req else ""
        classifier_p = (req.metadata.get("classifier_score") if req else None) or 0
        evidence = sorted(
            d.get("evidence") or [],
            key=lambda e: (e.get("judgment") or {}).get("llm_confidence") or 0,
            reverse=True,
        )
        top_ev = ""
        if evidence:
            top_text = _truncate(evidence[0].get("text") or "", 200)
            top_conf = (evidence[0].get("judgment") or {}).get("llm_confidence") or 0
            top_ev = f"{top_text}  [conf={top_conf:.2f}]"

        if status == "CONFLICT":
            out.append({"status": status, "req": req_text,
                         "reason": "Rule verifier detected a numeric conflict.",
                         "evidence": top_ev})
        elif status == "MISSING" and classifier_p >= 0.9:
            out.append({"status": status, "req": req_text,
                         "reason": f"High-confidence requirement (p={classifier_p:.2f}) with no matching PMI step.",
                         "evidence": top_ev})
        elif status == "COVERED" and evidence:
            top_conf = (evidence[0].get("judgment") or {}).get("llm_confidence") or 0
            if top_conf < 0.5:
                out.append({"status": status, "req": req_text,
                             "reason": f"Weak match (conf={top_conf:.2f}). Human review recommended.",
                             "evidence": top_ev})
    return out


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--packages", nargs="+",
                    default=["pkg_0002", "pkg_0005", "pkg_0008", "pkg_0010"])
    ap.add_argument("--packages-dir", default="./data/packages")
    ap.add_argument("--out", default="./demo_report.xlsx")
    ap.add_argument("--threshold", type=float, default=0.7)
    args = ap.parse_args()

    config = CoverageConfig()
    config.requirement_extraction = "model"
    config.requirement_model.threshold = args.threshold
    config.embedding.backend = "bow"
    config.llm.enabled = True
    config.llm.backend = "cross_encoder"

    pipeline = CoverageAnalysisPipeline(config)
    req_builder = RequirementBuilder(config)
    unit_builder = CoverageUnitBuilder()

    wb = Workbook()
    # Overwrite default sheet at the end (as Summary)
    default_sheet = wb.active
    default_sheet.title = "_tmp"

    per_package: Dict[str, Dict] = {}

    for pkg_name in args.packages:
        pkg_path = Path(args.packages_dir) / f"{pkg_name}.json"
        if not pkg_path.exists():
            print(f"[skip] {pkg_path} not found", file=sys.stderr)
            continue
        pkg = json.loads(pkg_path.read_text(encoding="utf-8"))
        request = _chunks_to_prepared(pkg)

        source_artifact = next(
            (d["prepared_artifact"] for d in request["documents"] if d["doc_role"] == "tz"),
            None,
        )
        if not source_artifact:
            continue

        requirements = req_builder.build(source_artifact)
        target_artifacts = [d["prepared_artifact"] for d in request["documents"]
                            if d["doc_role"] in ("pmi", "pz")]
        for art in target_artifacts:
            unit_builder.build(art)

        print(f"[info] running pipeline on {pkg_name} ({len(requirements)} requirements)")
        report = pipeline.run(request)
        report_d = report.model_dump() if hasattr(report, "model_dump") else report
        results = report_d.get("requirement_results") or []
        counts = Counter(_short_status(r.get("status")) for r in results)

        findings = _collect_findings(results, requirements)
        per_package[pkg_name] = {"counts": counts, "findings": findings}

        ws = wb.create_sheet(title=pkg_name)
        _write_package_sheet(ws, pkg_name, requirements, results,
                             report_d.get("pair_judgments") or [])

    # Summary + Findings sheets
    summary_ws = wb.create_sheet(title="Summary", index=0)
    _write_summary_sheet(summary_ws, per_package)
    findings_ws = wb.create_sheet(title="Findings", index=1)
    _write_findings_sheet(findings_ws, per_package)

    # Drop the temporary sheet
    del wb["_tmp"]

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f"[done] wrote {out_path}")


if __name__ == "__main__":
    main()
